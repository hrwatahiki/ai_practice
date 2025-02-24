import openpyxl.reader.excel
import sqlite3

from program.config import DB_PATH

target_columns = [ { "index": col, "col_type": ("str" if col in [1,2,3,4,62] else "float")} for col in range(1, 63) if col not in [15,18,33] ]

def get_value(value: str) -> str | float | None:
    if value == "-":
        return 0.0
    if value.startswith("(") and value.endswith(")"):
        value = value[1:-1]
    if value == "Tr":
        return 0
    try:
        return float(value)
    except:
        pass
    return None


def load_food(data_filename: str):
    # print(target_columns)
    try:
        wb = openpyxl.reader.excel.load_workbook(data_filename)
        sheet = wb["表全体"]
        row = 13
        records = []
        while True:
            if not sheet.cell(row, 4).value:
                break
            record = []
            for columns in target_columns:
                if columns["col_type"] == "str":
                    value = str(sheet.cell(row, columns["index"]).value)
                else:
                    value = get_value(str(sheet.cell(row, columns["index"]).value))
                record.append(value)
            # print(record)
            records.append(record)
            row += 1
    finally:
        wb.close()
    return records


def import_food(db_filename:str, records):
    with sqlite3.connect(db_filename) as conn:
        conn.execute("DELETE FROM '食品'")
        conn.executemany(f"INSERT INTO '食品' VALUES({"?,"*(len(target_columns)-1)}?)", records)

if __name__ == "__main__":
    records = load_food(r"D:\Users\Other\venv\ai\data\食品.xlsx")
    import_food(DB_PATH, records)